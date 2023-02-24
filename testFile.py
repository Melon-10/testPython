import openpyxl
def read_excel(file_path,sheet_name):
    # 加载整个文件到文档对象
    wb = openpyxl.load_workbook(file_path)
    sheet_data = wb[sheet_name]
    lines_count = sheet_data.max_row  # 获取总行数
    print(lines_count)
    cols_count = sheet_data.max_column  # 获取总列数
    print(cols_count)
    # 注意openxl读取行喝列时都是从1开始
    # 第一行我们是表头所以我们便利时不读
    data = []
    for l in range(2, lines_count+1):
        line = []  # 储存当前行数据
        for c in range(1, cols_count+1):
            cell_data = sheet_data.cell(l, c).value
            if cell_data is None:
                cell_data=""
            line.append(cell_data)  # 将当前单元格的数据添加到列表
        data.append(line)           # 将列表中的数据添加到data列表
    return data
if __name__ == '__main__':
    print(read_excel('test_data.xlsx', '登录接口数据'))